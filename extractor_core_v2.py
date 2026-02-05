"""
Customs Data Extractor V2 - Core Module
Supports both Export (TKX) and Import (TKN) declarations
"""

import xlrd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import re
from enum import Enum
from abc import ABC, abstractmethod
from typing import List, Dict, Optional, Callable, Union
from pathlib import Path


class DeclarationType(Enum):
    """Type of customs declaration"""
    EXPORT = "export"
    IMPORT = "import"


class ExtractionProgress:
    """Progress tracking for extraction process"""
    
    def __init__(self):
        self.total_steps = 0
        self.current_step = 0
        self.status_message = ""
        self.is_complete = False
        self.has_error = False
        self.error_message = ""
    
    @property
    def progress_percent(self) -> int:
        """Get progress as percentage (0-100)"""
        if self.total_steps == 0:
            return 0
        return int((self.current_step / self.total_steps) * 100)


class BaseExtractor(ABC):
    """Base class for customs declaration extractors"""
    
    @abstractmethod
    def get_sheet_aliases(self) -> List[str]:
        """Get list of valid sheet names/aliases to process"""
        pass
    
    @abstractmethod
    def extract_block_data(self, start_row: int) -> Optional[Dict[str, any]]:
        """Extract data from a single block"""
        pass
    
    def __init__(self, input_file: str, progress_callback: Optional[Callable] = None):
        """Initialize extractor"""
        self.input_file = input_file
        self.progress_callback = progress_callback
        self.progress = ExtractionProgress()
        self.workbook = None
        self.sheet = None
        self.data_blocks = []
        self.found_sheet_name = None
        
        # Detect file format
        self.file_ext = Path(input_file).suffix.lower()
        self.is_xls = (self.file_ext == '.xls')
    
    def _update_progress(self, step: int, message: str):
        """Update progress and call callback"""
        self.progress.current_step = step
        self.progress.status_message = message
        
        if self.progress_callback:
            self.progress_callback(self.progress)
    
    def get_cell_value(self, row: int, col: int):
        """Get cell value - works for both xlrd and openpyxl"""
        if self.is_xls:
            return self.sheet.cell_value(row, col)
        else:
            return self.sheet.cell(row, col).value
    
    @staticmethod
    def format_number(value) -> Union[float, int, str]:
        """
        Format number from Vietnamese or US format to Excel number.
        - 1.000,50 -> 1000.5
        - 1,000.50 -> 1000.5
        - 100,50 -> 100.5
        - 100.50 -> 100.5
        - 10.000 -> 10000
        """
        if value is None or value == "":
            return ""
        
        value_str = str(value).strip()
        if not value_str:
            return ""
        
        # Already a number
        if isinstance(value, (int, float)):
            return value

        try:
            # Check if we have both dot and comma
            if '.' in value_str and ',' in value_str:
                last_dot = value_str.rfind('.')
                last_comma = value_str.rfind(',')
                
                if last_comma > last_dot:
                    # VN Style: 1.000,50 -> Remove dots, replace comma with dot
                    clean_str = value_str.replace('.', '').replace(',', '.')
                else:
                    # US Style: 1,000.50 -> Remove commas, keep dot
                    clean_str = value_str.replace(',', '')
            
            elif ',' in value_str:
                # Only comma: 100,50 -> 100.5(decimals)
                # Assumption: In text context here, comma is usually decimal separator
                clean_str = value_str.replace(',', '.')
                
            elif '.' in value_str:
                # Only dot: 
                # 1.000.000 -> 1000000
                # 10.000 -> 10000
                # 100.50 -> 100.5
                
                # If multiple dots, definitely thousands separator
                if value_str.count('.') > 1:
                    clean_str = value_str.replace('.', '')
                else:
                    # Single dot
                    # Heuristic: If followed by exactly 3 digits, assume thousands
                    # e.g. 10.000 -> 10000
                    # e.g. 10.5 -> 10.5
                    if re.search(r'\.\d{3}$', value_str):
                        clean_str = value_str.replace('.', '')
                    else:
                        clean_str = value_str

            else:
                clean_str = value_str

            # Convert to number
            if '.' in clean_str:
                return float(clean_str)
            return int(clean_str)

        except ValueError:
            return value_str
    
    def find_declaration_id(self) -> str:
        """
        Scan the first 50 rows to find declaration ID.
        Export ID: Starts with '30', 12 digits.
        Import ID: Starts with '10', 12 digits.
        Returns empty string if not found.
        """
        if not self.sheet:
            return ""
            
        pattern = re.compile(r'\b(10\d{10}|30\d{10})\b')
        
        # Determine max rows to scan (usually info is in header)
        if self.is_xls:
            max_scan = min(50, self.sheet.nrows)
        else:
            max_scan = min(50, self.sheet.max_row)
            
        for r in range(max_scan):
            if self.is_xls:
                row_values = self.sheet.row_values(r)
            else:
                # openpyxl is 1-based for rows
                row_values = [cell.value for cell in self.sheet[r + 1]]
                
            for val in row_values:
                if val:
                    val_str = str(val).strip()
                    match = pattern.search(val_str)
                    if match:
                        return match.group(1)
        return ""

    def load_workbook(self) -> bool:
        """Load Excel workbook and find matching sheet"""
        try:
            self._update_progress(0, f"Đang mở file: {Path(self.input_file).name}")
            
            aliases = [a.lower() for a in self.get_sheet_aliases()]
            
            if self.is_xls:
                self.workbook = xlrd.open_workbook(self.input_file)
                sheet_names = self.workbook.sheet_names()
                
                # Find matching sheet
                target_sheet = None
                for name in sheet_names:
                    if name.lower() in aliases:
                        target_sheet = name
                        break
                
                if not target_sheet:
                    raise Exception(f"Không tìm thấy sheet nào phù hợp. Các tên hỗ trợ: {', '.join(self.get_sheet_aliases())}")

                self.found_sheet_name = target_sheet
                self.sheet = self.workbook.sheet_by_name(target_sheet)
                nrows = self.sheet.nrows
                ncols = self.sheet.ncols
            else:
                self.workbook = load_workbook(self.input_file, data_only=True)
                sheet_names = self.workbook.sheetnames
                
                # Find matching sheet
                target_sheet = None
                for name in sheet_names:
                    if name.lower() in aliases:
                        target_sheet = name
                        break
                
                if not target_sheet:
                    raise Exception(f"Không tìm thấy sheet nào phù hợp. Các tên hỗ trợ: {', '.join(self.get_sheet_aliases())}")

                self.found_sheet_name = target_sheet
                self.sheet = self.workbook[target_sheet]
                nrows = self.sheet.max_row
                ncols = self.sheet.max_column
            
            self._update_progress(1, f"✓ Đã load sheet '{self.found_sheet_name}' ({nrows} hàng, {ncols} cột)")
            return True
        except Exception as e:
            self.progress.has_error = True
            self.progress.error_message = f"Lỗi khi mở file: {str(e)}"
            if self.progress_callback:
                self.progress_callback(self.progress)
            return False
    
    def get_preview_data(self) -> List[Dict[str, str]]:
        """Get preview of data blocks for display"""
        preview = []
        for idx, block_start in enumerate(self.data_blocks[:20], 1):
            data = self.extract_block_data(block_start)
            if data:
                desc = data.get('description', '')
                origin = data.get('origin', '')
                if origin:
                    desc_display = f"{desc} ({origin})"
                else:
                    desc_display = desc
                
                preview.append({
                    'index': idx,
                    'hs_code': data.get('hs_code', ''),
                    'description': desc_display[:100] + '...' if len(desc_display) > 100 else desc_display,
                    'qty1': str(data.get('qty1', '')),
                    'unit1': data.get('unit1', '')
                })
        return preview
    
    def create_output_file(self, output_file: str) -> bool:
        """Create Excel output file with extracted data"""
        try:
            self._update_progress(4, f"Đang tạo file output...")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "data"
            
            # Header
            # STT -> Mô tả -> HS Code -> Xuất xứ -> So luong 1 -> Don vi 1 -> ...
            headers = [
                "STT",
                "Mô tả hàng hóa", 
                "HS code", 
                "Xuất xứ",
                "Số lượng (1)", 
                "Đơn vị 1", 
                "Số lượng (2)", 
                "Đơn vị 2",
                "Đơn giá hóa đơn",
                "Trị giá hóa đơn"
            ]
            ws.append(headers)
            
            # Format header
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Write data
            total_blocks = len(self.data_blocks)
            for idx, block_start in enumerate(self.data_blocks, 1):
                self._update_progress(4 + idx, f"Đang ghi dữ liệu khối {idx}/{total_blocks}...")
                
                data = self.extract_block_data(block_start)
                if data:
                    row = [
                        idx, # STT
                        data.get('description', ''),
                        data.get('hs_code', ''),
                        data.get('origin', ''),
                        data.get('qty1', ''),
                        data.get('unit1', ''),
                        data.get('qty2', ''),
                        data.get('unit2', ''),
                        data.get('unit_price', ''),
                        data.get('invoice_value', '')
                    ]
                    ws.append(row)
                    
                    # Store row index to apply formatting later (1-based, +1 for header)
                    current_excel_row = idx + 1
                    
                    price_cell = ws.cell(row=current_excel_row, column=9)
                    val_price = data.get('unit_price')
                    
                    if isinstance(val_price, (int, float)):
                        # Conditional formatting: if integer -> no decimals, else 2 decimals
                        if val_price % 1 == 0:
                            price_cell.number_format = '#,##0'
                            
                            # Also ensure the value stored is integer if it looks like float (100.0 -> 100)
                            price_cell.value = int(val_price)
                        else:
                            price_cell.number_format = '#,##0.00'
                        
                    value_cell = ws.cell(row=current_excel_row, column=10)
                    val_invoice = data.get('invoice_value')
                    
                    if isinstance(val_invoice, (int, float)):
                        if val_invoice % 1 == 0:
                            value_cell.number_format = '#,##0'
                            value_cell.value = int(val_invoice)
                        else:
                            value_cell.number_format = '#,##0.00'

            
            # Format columns width
            ws.column_dimensions['A'].width = 8   # STT
            ws.column_dimensions['B'].width = 70  # Mo ta
            ws.column_dimensions['C'].width = 15  # HS Code
            ws.column_dimensions['D'].width = 15  # Xuat xu
            ws.column_dimensions['E'].width = 12  # QC1
            ws.column_dimensions['F'].width = 15  # Unit1
            ws.column_dimensions['G'].width = 12  # QC2
            ws.column_dimensions['H'].width = 15  # Unit2
            ws.column_dimensions['I'].width = 18  # Don gia
            ws.column_dimensions['J'].width = 18  # Tri gia
            
            # Borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=1, max_row=len(self.data_blocks) + 1, 
                                   min_col=1, max_col=10):
                for cell in row:
                    cell.border = thin_border
                    # Center align STT and Units
                    if cell.col_idx in [1, 5, 6, 7, 8]:
                        cell.alignment = Alignment(horizontal="center", vertical="top")
                    # Right align for Numbers
                    elif cell.col_idx in [9, 10]:
                        cell.alignment = Alignment(horizontal="right", vertical="top")
                    else:
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Save
            wb.save(output_file)
            self._update_progress(4 + total_blocks + 1, f"✓ Đã lưu file: {Path(output_file).name}")
            return True
            
        except Exception as e:
            self.progress.has_error = True
            self.progress.error_message = f"Lỗi khi tạo file: {str(e)}"
            if self.progress_callback:
                self.progress_callback(self.progress)
            return False
    
    def run(self, output_file: str) -> bool:
        """Run complete extraction process"""
        self.progress.total_steps = 5 + len(self.data_blocks) if hasattr(self, 'data_blocks') else 10
        
        if not self.load_workbook():
            return False
        
        num_blocks = self.find_data_blocks()
        if num_blocks == 0:
            self.progress.has_error = True
            self.progress.error_message = "Không tìm thấy dữ liệu nào!"
            if self.progress_callback:
                self.progress_callback(self.progress)
            return False
        
        self.progress.total_steps = 5 + num_blocks
        
        if not self.create_output_file(output_file):
            return False
        
        self.progress.is_complete = True
        self.progress.current_step = self.progress.total_steps
        self.progress.status_message = f"✓ Hoàn thành! Đã trích xuất {num_blocks} khối dữ liệu"
        if self.progress_callback:
            self.progress_callback(self.progress)
        
        return True


class ExportExtractor(BaseExtractor):
    """Extractor for Export declarations (TKX)"""
    
    # Column definitions (0-based)
    COL_LABEL = 2
    COL_VALUE = 5       # Column F
    COL_QTY_VALUE = 16  # Column Q
    COL_UNIT = 24       # Column Y
    COL_INVOICE_PRICE = 17  # Column R
    
    # Offsets
    OFFSET_DESCRIPTION = 1
    OFFSET_QTY1 = 4
    OFFSET_QTY2 = 5
    OFFSET_INVOICE = 6
    
    def get_sheet_aliases(self) -> List[str]:
        return ['tkx', 'tokhaixuat02']
    
    @staticmethod
    def extract_origin(description: str) -> tuple:
        """Extract country of origin from description pattern #&XX"""
        if not description:
            return "", ""
        
        pattern = r'#&([A-Z]{2,})\s*$'
        match = re.search(pattern, description)
        
        if match:
            origin = match.group(1)
            return description, origin
        
        return description, ""
    
    def find_data_blocks(self) -> int:
        """Find all data blocks in export sheet"""
        try:
            self._update_progress(2, "Đang tìm kiếm các khối dữ liệu...")
            self.data_blocks = []
            
            if self.is_xls:
                max_rows = self.sheet.nrows
            else:
                max_rows = self.sheet.max_row
            
            for row_idx in range(max_rows):
                try:
                    if self.is_xls:
                        cell_value = self.get_cell_value(row_idx, self.COL_LABEL)
                    else:
                        cell_value = self.get_cell_value(row_idx + 1, self.COL_LABEL + 1)
                    
                    if cell_value and "Mã số hàng hóa" in str(cell_value):
                        if self.is_xls:
                            hs_code = self.get_cell_value(row_idx, self.COL_VALUE)
                        else:
                            hs_code = self.get_cell_value(row_idx + 1, self.COL_VALUE + 1)
                        
                        hs_code_str = str(hs_code).strip() if hs_code else ""
                        
                        if isinstance(hs_code, (int, float)):
                            hs_code_str = str(int(hs_code))
                        
                        if hs_code_str.isdigit() and len(hs_code_str) == 8:
                            if self.is_xls:
                                self.data_blocks.append(row_idx)
                            else:
                                self.data_blocks.append(row_idx + 1)
                except:
                    continue
            
            self._update_progress(3, f"✓ Tìm thấy {len(self.data_blocks)} khối dữ liệu")
            return len(self.data_blocks)
            
        except Exception as e:
            self.progress.has_error = True
            self.progress.error_message = f"Lỗi khi tìm dữ liệu: {str(e)}"
            if self.progress_callback:
                self.progress_callback(self.progress)
            return 0
    
    def extract_block_data(self, start_row: int) -> Optional[Dict[str, any]]:
        """Extract data from export declaration block"""
        try:
            if self.is_xls:
                hs_code = self.get_cell_value(start_row, self.COL_VALUE)
                description_raw = self.get_cell_value(start_row + self.OFFSET_DESCRIPTION, self.COL_VALUE)
                qty1 = self.get_cell_value(start_row + self.OFFSET_QTY1, self.COL_QTY_VALUE)
                unit1 = self.get_cell_value(start_row + self.OFFSET_QTY1, self.COL_UNIT)
                qty2 = self.get_cell_value(start_row + self.OFFSET_QTY2, self.COL_QTY_VALUE)
                unit2 = self.get_cell_value(start_row + self.OFFSET_QTY2, self.COL_UNIT)
                invoice_value = self.get_cell_value(start_row + self.OFFSET_INVOICE, self.COL_VALUE)
                unit_price = self.get_cell_value(start_row + self.OFFSET_INVOICE, self.COL_INVOICE_PRICE)
            else:
                hs_code = self.get_cell_value(start_row, self.COL_VALUE + 1)
                description_raw = self.get_cell_value(start_row + self.OFFSET_DESCRIPTION, self.COL_VALUE + 1)
                qty1 = self.get_cell_value(start_row + self.OFFSET_QTY1, self.COL_QTY_VALUE + 1)
                unit1 = self.get_cell_value(start_row + self.OFFSET_QTY1, self.COL_UNIT + 1)
                qty2 = self.get_cell_value(start_row + self.OFFSET_QTY2, self.COL_QTY_VALUE + 1)
                unit2 = self.get_cell_value(start_row + self.OFFSET_QTY2, self.COL_UNIT + 1)
                invoice_value = self.get_cell_value(start_row + self.OFFSET_INVOICE, self.COL_VALUE + 1)
                unit_price = self.get_cell_value(start_row + self.OFFSET_INVOICE, self.COL_INVOICE_PRICE + 1)
            
            data = {}
            
            # HS code
            if isinstance(hs_code, (int, float)):
                data['hs_code'] = str(int(hs_code))
            else:
                data['hs_code'] = str(hs_code).strip() if hs_code else ""
            
            # Description + origin
            description_str = str(description_raw).strip() if description_raw else ""
            original_desc, origin = self.extract_origin(description_str)
            data['description'] = original_desc
            data['origin'] = origin
            
            # Format numbers (returns float/int/str)
            data['qty1'] = self.format_number(qty1)
            data['unit1'] = str(unit1).strip() if unit1 else ""
            data['qty2'] = self.format_number(qty2)
            data['unit2'] = str(unit2).strip() if unit2 else ""
            data['invoice_value'] = self.format_number(invoice_value)
            data['unit_price'] = self.format_number(unit_price)
            
            return data
            
        except Exception as e:
            return None


class ImportExtractor(BaseExtractor):
    """Extractor for Import declarations (TKN)"""
    
    # Column definitions (0-based)
    COL_LABEL = 2       # Column C
    COL_VALUE = 6       # Column G
    COL_QTY_VALUE = 21  # Column V
    COL_UNIT = 30       # Column AF (31 in 1-based = PCE)
    COL_INVOICE_VALUE = 8   # Column I
    COL_ORIGIN = 23     # Column X
    
    # Offsets
    OFFSET_DESCRIPTION = 1
    OFFSET_QTY1 = 4
    OFFSET_QTY2 = 5
    OFFSET_INVOICE = 6
    OFFSET_ORIGIN = 11  # Row N+11 for origin
    
    def get_sheet_aliases(self) -> List[str]:
        return ['tkn', 'tokhainhap2', 'tokhainhap02', 'tờ khai nhập']
    
    def find_data_blocks(self) -> int:
        """Find all data blocks in import sheet"""
        try:
            self._update_progress(2, "Đang tìm kiếm các khối dữ liệu...")
            self.data_blocks = []
            
            if self.is_xls:
                max_rows = self.sheet.nrows
            else:
                max_rows = self.sheet.max_row
            
            for row_idx in range(max_rows):
                try:
                    if self.is_xls:
                        cell_value = self.get_cell_value(row_idx, self.COL_LABEL)
                    else:
                        cell_value = self.get_cell_value(row_idx + 1, self.COL_LABEL + 1)
                    
                    if cell_value and "Mã số hàng hóa" in str(cell_value):
                        if self.is_xls:
                            hs_code = self.get_cell_value(row_idx, self.COL_VALUE)
                        else:
                            hs_code = self.get_cell_value(row_idx + 1, self.COL_VALUE + 1)
                        
                        hs_code_str = str(hs_code).strip() if hs_code else ""
                        
                        if isinstance(hs_code, (int, float)):
                            hs_code_str = str(int(hs_code))
                        
                        if hs_code_str.isdigit() and len(hs_code_str) == 8:
                            if self.is_xls:
                                self.data_blocks.append(row_idx)
                            else:
                                self.data_blocks.append(row_idx + 1)
                except:
                    continue
            
            self._update_progress(3, f"✓ Tìm thấy {len(self.data_blocks)} khối dữ liệu")
            return len(self.data_blocks)
            
        except Exception as e:
            self.progress.has_error = True
            self.progress.error_message = f"Lỗi khi tìm dữ liệu: {str(e)}"
            if self.progress_callback:
                self.progress_callback(self.progress)
            return 0
    
    def extract_block_data(self, start_row: int) -> Optional[Dict[str, any]]:
        """Extract data from import declaration block"""
        try:
            if self.is_xls:
                hs_code = self.get_cell_value(start_row, self.COL_VALUE)
                description_raw = self.get_cell_value(start_row + self.OFFSET_DESCRIPTION, self.COL_VALUE)
                qty1 = self.get_cell_value(start_row + self.OFFSET_QTY1, self.COL_QTY_VALUE)
                unit1 = self.get_cell_value(start_row + self.OFFSET_QTY1, self.COL_UNIT)
                qty2 = self.get_cell_value(start_row + self.OFFSET_QTY2, self.COL_QTY_VALUE)
                unit2 = self.get_cell_value(start_row + self.OFFSET_QTY2, self.COL_UNIT)
                invoice_value = self.get_cell_value(start_row + self.OFFSET_INVOICE, self.COL_INVOICE_VALUE)
                unit_price = self.get_cell_value(start_row + self.OFFSET_INVOICE, self.COL_QTY_VALUE)
                origin = self.get_cell_value(start_row + self.OFFSET_ORIGIN, self.COL_ORIGIN)
            else:
                hs_code = self.get_cell_value(start_row, self.COL_VALUE + 1)
                description_raw = self.get_cell_value(start_row + self.OFFSET_DESCRIPTION, self.COL_VALUE + 1)
                qty1 = self.get_cell_value(start_row + self.OFFSET_QTY1, self.COL_QTY_VALUE + 1)
                unit1 = self.get_cell_value(start_row + self.OFFSET_QTY1, self.COL_UNIT + 1)
                qty2 = self.get_cell_value(start_row + self.OFFSET_QTY2, self.COL_QTY_VALUE + 1)
                unit2 = self.get_cell_value(start_row + self.OFFSET_QTY2, self.COL_UNIT + 1)
                invoice_value = self.get_cell_value(start_row + self.OFFSET_INVOICE, self.COL_INVOICE_VALUE + 1)
                unit_price = self.get_cell_value(start_row + self.OFFSET_INVOICE, self.COL_QTY_VALUE + 1)
                origin = self.get_cell_value(start_row + self.OFFSET_ORIGIN, self.COL_ORIGIN + 1)
            
            data = {}
            
            # HS code
            if isinstance(hs_code, (int, float)):
                data['hs_code'] = str(int(hs_code))
            else:
                data['hs_code'] = str(hs_code).strip() if hs_code else ""
            
            # Description (no origin extraction from text)
            data['description'] = str(description_raw).strip() if description_raw else ""
            
            # Origin from dedicated row/cell
            data['origin'] = str(origin).strip() if origin else ""
            
            # Format numbers (returns float/int/str)
            data['qty1'] = self.format_number(qty1)
            data['unit1'] = str(unit1).strip() if unit1 else ""
            data['qty2'] = self.format_number(qty2)
            data['unit2'] = str(unit2).strip() if unit2 else ""
            data['invoice_value'] = self.format_number(invoice_value)
            data['unit_price'] = self.format_number(unit_price)
            
            return data
            
        except Exception as e:
            return None
